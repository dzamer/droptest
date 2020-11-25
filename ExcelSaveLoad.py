from openpyxl import load_workbook, Workbook
import datetime


class ExcelOperation:
    timeData = []
    upperSensorData = []
    lowerSensorData = []
    accelerometerXData = []
    accelerometerYData = []
    accelerometerZData = []

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def load_from_excel(self, workbook, worksheet):
        wb = load_workbook('{}.xlsx'.format(workbook))
        print(wb.sheetnames)
        sheet = wb['{}'.format(worksheet)]
        max_rows = sheet.max_row
        for var in range(1, max_rows):
            self.timeData.append(float(sheet.cell(row=var, column=1).value))
            self.upperSensorData.append(float(sheet.cell(row=var, column=2).value))
            self.lowerSensorData.append(float(sheet.cell(row=var, column=3).value))
            self.accelerometerXData.append(float(sheet.cell(row=var, column=4).value))
            self.accelerometerYData.append(float(sheet.cell(row=var, column=5).value))
            self.accelerometerZData.append(float(sheet.cell(row=var, column=6).value))
        print("woorksheet loading completed")

    def save_to_excel(self, workbookName, refDistance, refTime, fallTime, impactForceMethod1, impactForceMethod2):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = datetime.datetime.now()
        worksheet['A2'] = "Time"
        worksheet['B2'] = "Upper Sensor"
        worksheet['C2'] = "Lower Sensor"
        worksheet['D2'] = "Acce X"
        worksheet['E2'] = "Acce Y"
        worksheet['F2'] = "Acce Z"
        worksheet['G2'] = "Reflection Travel: {}".format(refDistance)
        worksheet['H2'] = "Reflection Time: {}".format(refTime)
        worksheet['I2'] = "Fall time:{}".format(fallTime)
        worksheet['J2'] = "Force from distance: {}".format(impactForceMethod1)
        worksheet['K2'] = "Force from time: {}".format(impactForceMethod2)
        for var in range(0, len(self.timeData)):
            worksheet.cell(var + 3, 1, self.timeData[var])
            worksheet.cell(var + 3, 2, self.upperSensorData[var])
            worksheet.cell(var + 3, 3, self.lowerSensorData[var])
            worksheet.cell(var + 3, 4, self.accelerometerXData[var])
            worksheet.cell(var + 3, 5, self.accelerometerYData[var])
            worksheet.cell(var + 3, 6, self.accelerometerZData[var])
        workbook.save("{}.xlsx".format(workbookName))
