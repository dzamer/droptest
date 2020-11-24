from PyQt5 import QtWidgets, uic, QtGui
from PyQt5.QtSerialPort import QSerialPort, QSerialPortInfo
from PyQt5.QtCore import QObject, QIODevice, pyqtSignal
import pyqtgraph as pg
import sys
import threading
import math
import datetime
from openpyxl import load_workbook, Workbook


class SerialPort(QObject):
    mySerialPort = QSerialPort()
    devicePortInfo = None
    readedData = ""
    dataToWrite = ""
    deviceIsFound = False
    deviceIsConnected = False
    waitForReadyReadValue = 200
    productIdentifier = 29987  # product ID to be set for device if other will be used (use serachForDevices if needed)
    vendorIdentifier = 6790  # vendor ID to be set for device if other will be used (use serachForDevices if needed)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.search_for_devices()

    # use serachForDevices if you need to get vendor and product ID for findDevice function
    def search_for_devices(self):
        device = QSerialPortInfo
        availablePort = device.availablePorts()

        i = 0
        for portInfo in availablePort:
            print("Device: {0}, product ID: {1}, vendor ID: {2}, Port name: {3}".format(i,
                                                                                        portInfo.productIdentifier(),
                                                                                        portInfo.vendorIdentifier(),
                                                                                        portInfo.portName()))
            i += 1

    def find_device(self):
        info = QSerialPortInfo
        availablePorts = info.availablePorts()

        for portInfo in availablePorts:
            print(type(portInfo.productIdentifier()))
            if portInfo.productIdentifier() == self.productIdentifier and portInfo.vendorIdentifier() \
                    == self.vendorIdentifier:
                self.devicePortInfo = portInfo
                self.deviceIsFound = True
                print("Device found at port: {0}".format(portInfo.portName))
                break

    def open_serial_port(self):
        try:
            self.mySerialPort.setPortName(str(self.devicePortInfo.portName()))
            self.mySerialPort.open(QIODevice.ReadWrite)
            self.mySerialPort.setBaudRate(2000000)
            self.mySerialPort.setDataBits(self.mySerialPort.Data8)
            print("Opening port {0}".format(self.mySerialPort.portName()))
            self.mySerialPort.setDataTerminalReady(True)
            self.deviceIsConnected = True
            print(self.mySerialPort.isOpen())
            print(self.mySerialPort.portName())
            self.mySerialPort.startTransaction()
            print(self.mySerialPort.isDataTerminalReady())
            print(self.mySerialPort.isReadable())
            print(self.mySerialPort.isWritable())
            return self.mySerialPort
        except Exception as e:
            print("open_serial_port: ", e)
            return None

    def read_serial(self):
        if self.deviceIsConnected:
            self.mySerialPort.waitForReadyRead(self.waitForReadyReadValue)
            temp = str(self.mySerialPort.readLineData(1024))
            temp2 = temp.split("'")
            temp3 = temp2[1].split("\\r")
            self.readedData = temp3[0]
        else:
            print("Device not connected")
            return

    def write_serial(self):
        if self.deviceIsConnected:
            dataEncoded = self.dataToWrite.encode()
            self.mySerialPort.write(dataEncoded)
            print("You enter {0} to serial.".format(self.dataToWrite))
        else:
            print("Device not connected")


class MainWindow(QtWidgets.QWidget):
    signal1 = pyqtSignal(str)
    readingLines = False
    readingLoop = True
    readedDataFromArduino = []
    clickedButtonData = ""
    fullReadedDate = ""
    fullTimeDate = []
    fullLoadSensorDate = []
    fullUpperSensorDate = []
    fullLowerSensorDate = []
    fullAccelerometerXDate = []
    fullAccelerometerYDate = []
    fullAccelerometerZDate = []
    height = 1.06
    upperSensor = 0.0
    lowerSensor = 0.0
    fallTime = 0.0
    loadMax = 0.0
    acceMax = 0.0
    acceMin = 0.0
    loadInTime = 0.0
    acceInTime = 0.0
    acceLowInTime = 0.0
    brakingTime = 0.0
    averageAcce = 0.0
    velocity = 0.0
    reflectionTime = 0.0
    reflectionDistance = 0.0
    reflectionAcceleration = 0.0
    upperIndex = 0.0
    lowerIndex = 0.0
    wnet = 0.0
    mass = 2.0
    impactForceMethod1 = 0.0
    impactForceMethod2 = 0.0

    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        uic.loadUi('DropTest.ui', self)
        self.setWindowTitle('Drop Test Window')
        self.Arduino = SerialPort()
        self.statusBar = QtWidgets.QStatusBar()
        # przyciski wszystkie sa włączone po to żeby szło aktualnie programem operować bez arduino
        # navigate buttons
        self.Button_exit.clicked.connect(self.close)
        self.SearchForDevices.clicked.connect(self.Arduino.search_for_devices)
        self.ConnectButton.clicked.connect(self.arduino_connection)
        # buttons to set accelerometer range
        self.Button_range_2G.setDisabled(True)
        self.Button_range_2G.clicked.connect(self.button_clicked_2G)
        self.Button_range_4G.setDisabled(True)
        self.Button_range_4G.clicked.connect(self.button_clicked_4G)
        self.Button_range_8G.setDisabled(True)
        self.Button_range_8G.clicked.connect(self.button_clicked_8G)
        self.Button_range_16G.setDisabled(True)
        self.Button_range_16G.clicked.connect(self.button_clicked_16G)
        # buttons to set accelerometer data rate
        self.Button_data_rate_1Hz.setDisabled(True)
        self.Button_data_rate_1Hz.clicked.connect(self.button_clicked_1Hz)
        self.Button_data_rate_10Hz.setDisabled(True)
        self.Button_data_rate_10Hz.clicked.connect(self.button_clicked_10Hz)
        self.Button_data_rate_25Hz.setDisabled(True)
        self.Button_data_rate_25Hz.clicked.connect(self.button_clicked_25Hz)
        self.Button_data_rate_50Hz.setDisabled(True)
        self.Button_data_rate_50Hz.clicked.connect(self.button_clicked_50Hz)
        self.Button_data_rate_100Hz.setDisabled(True)
        self.Button_data_rate_100Hz.clicked.connect(self.button_clicked_100Hz)
        self.Button_data_rate_200Hz.setDisabled(True)
        self.Button_data_rate_200Hz.clicked.connect(self.button_clicked_200Hz)
        self.Button_data_rate_400Hz.setDisabled(True)
        self.Button_data_rate_400Hz.clicked.connect(self.button_clicked_400Hz)
        self.pushButton_data_rate_5KHz_lp.setDisabled(True)
        self.pushButton_data_rate_5KHz_lp.clicked.connect(self.button_clicked_5KHz)
        self.pushButton_data_rate_16KHz_lp.setDisabled(True)
        self.pushButton_data_rate_16KHz_lp.clicked.connect(self.button_clicked_16KHz)
        self.pushButton_data_rate_power_down.setDisabled(True)
        self.pushButton_data_rate_power_down.clicked.connect(self.button_clicked_power_down)
        self.Button_start_test.setDisabled(True)
        self.Button_start_test.clicked.connect(self.button_start)
        # temporary buttons
        self.Button_readData.setDisabled(True)
        self.Button_readData.clicked.connect(self.plot_generating)
        self.Button_reset.setDisabled(True)
        self.Button_reset.clicked.connect(self.arduino_reading_exit)
        self.plot_accelerometer(self.fullTimeDate, self.fullAccelerometerZDate, 'r', "Acce. Z")
        self.update()

    def read_date(self):
        self.Arduino.read_serial()
        if self.Arduino.readedData != "" and self.readingLines is False and \
                self.Arduino.readedData != "Droptest started..":
            self.textBrowser_main.append(self.Arduino.readedData)
        elif self.Arduino.readedData == "Droptest started..":
            self.textBrowser_main.append(self.Arduino.readedData)
            self.readingLines = True
            self.Arduino.waitForReadyReadValue = 50
        elif self.Arduino.readedData != "Droptest started.." and self.Arduino.readedData != "":
            self.readedDataFromArduino = (str(self.Arduino.readedData))
            print(self.readedDataFromArduino)
            temp2 = self.readedDataFromArduino.split(";")
            temp = temp2[0].split(",")
            self.fullTimeDate.append(float(temp[0]))
            self.fullUpperSensorDate.append(float(temp[1]))
            self.fullLowerSensorDate.append(float(temp[2]))
            self.fullAccelerometerXDate.append(float(temp[3]))
            self.fullAccelerometerYDate.append(float(temp[4]))
            self.fullAccelerometerZDate.append(float(temp[5]))

    def read_data_in_loop(self):
        self.readingLoop = True
        while self.readingLoop:
            self.read_date()

    def arduino_reading_exit(self):
        self.readingLoop = False
        self.textBrowser_main.append("Loading sensors data finished.")
        self.Button_readData.setDisabled(False)

    def arduino_connection(self):
        try:
            self.Arduino.find_device()
            self.Arduino.open_serial_port()
            thread = threading.Thread(target=self.read_data_in_loop)
            thread.start()
            if self.Arduino.mySerialPort.isOpen():
                self.range_button_enabling()
        except Exception as e:
            self.statusBar.showMessage("Arduino connection: {}".format(e), 2500)

    def range_button_enabling(self):
        self.Button_range_2G.setDisabled(False)
        self.Button_range_4G.setDisabled(False)
        self.Button_range_8G.setDisabled(False)
        self.Button_range_16G.setDisabled(False)

    def range_button_disabling(self):
        self.Button_range_2G.setDisabled(True)
        self.Button_range_4G.setDisabled(True)
        self.Button_range_8G.setDisabled(True)
        self.Button_range_16G.setDisabled(True)

    def data_rate_button_enabling(self):
        self.Button_data_rate_1Hz.setDisabled(False)
        self.Button_data_rate_10Hz.setDisabled(False)
        self.Button_data_rate_25Hz.setDisabled(False)
        self.Button_data_rate_50Hz.setDisabled(False)
        self.Button_data_rate_100Hz.setDisabled(False)
        self.Button_data_rate_200Hz.setDisabled(False)
        self.Button_data_rate_400Hz.setDisabled(False)
        self.pushButton_data_rate_5KHz_lp.setDisabled(False)
        self.pushButton_data_rate_16KHz_lp.setDisabled(False)
        self.pushButton_data_rate_power_down.setDisabled(False)

    def data_rate_button_disabling(self):
        self.Button_data_rate_1Hz.setDisabled(True)
        self.Button_data_rate_10Hz.setDisabled(True)
        self.Button_data_rate_25Hz.setDisabled(True)
        self.Button_data_rate_50Hz.setDisabled(True)
        self.Button_data_rate_100Hz.setDisabled(True)
        self.Button_data_rate_200Hz.setDisabled(True)
        self.Button_data_rate_400Hz.setDisabled(True)
        self.pushButton_data_rate_5KHz_lp.setDisabled(True)
        self.pushButton_data_rate_16KHz_lp.setDisabled(True)
        self.pushButton_data_rate_power_down.setDisabled(True)

    def button_clicked_2G(self):
        try:
            self.clickedButtonData = "12"
            self.data_rate_button_enabling()
            self.textBrowser_main.append("Accelerometer range set to: {0}".format(self.clickedButtonData))
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.range_button_disabling()
        except Exception as e:
            self.statusBar.showMessage("Range button: {}".format(e, 2500))

    def button_clicked_4G(self):
        try:
            self.clickedButtonData = "13"
            self.data_rate_button_enabling()
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.range_button_disabling()
        except Exception as e:
            self.statusBar.showMessage("Range button: {}".format(e, 2500))

    def button_clicked_8G(self):
        try:
            self.clickedButtonData = "14"
            self.data_rate_button_enabling()
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.range_button_disabling()
        except Exception as e:
            self.statusBar.showMessage("Range button: {}".format(e, 2500))

    def button_clicked_16G(self):
        try:
            self.clickedButtonData = "15"
            self.data_rate_button_enabling()
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.range_button_disabling()
        except Exception as e:
            self.statusBar.showMessage("Range button: {}".format(e, 2500))

    def button_clicked_1Hz(self):
        try:
            self.clickedButtonData = "21"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_10Hz(self):
        try:
            self.clickedButtonData = "22"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_25Hz(self):
        try:
            self.clickedButtonData = "23"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_50Hz(self):
        try:
            self.clickedButtonData = "24"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_100Hz(self):
        try:
            self.clickedButtonData = "25"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_200Hz(self):
        try:
            self.clickedButtonData = "26"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_400Hz(self):
        try:
            self.clickedButtonData = "27"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_5KHz(self):
        try:
            self.clickedButtonData = "28"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_16KHz(self):
        try:
            self.clickedButtonData = "29"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.range_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_clicked_power_down(self):
        try:
            self.clickedButtonData = "30"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.data_rate_button_disabling()
            self.Button_start_test.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def button_start(self):
        try:
            self.clickedButtonData = "41"
            self.Arduino.dataToWrite = self.clickedButtonData
            self.Arduino.write_serial()
            self.textBrowser_main.append("Loading sensors data...")
            self.Button_reset.setDisabled(False)
        except Exception as e:
            self.statusBar.showMessage("Data button: {}".format(e, 2500))

    def plot_generating(self):
        try:
            self.load_from_excel()
            tempTime = self.fullTimeDate[0]
            for var3 in range(0, len(self.fullTimeDate)):
                if self.fullTimeDate[var3] - self.fullTimeDate[0] == 0:
                    self.fullTimeDate[var3] = 0.0
                else:
                    self.fullTimeDate[var3] = (self.fullTimeDate[var3] - tempTime) / 1000
            # setting fall time and max values for load sensor and accelerometer
            self.fall_time()
            self.reflection_travel()
            self.impact_force_method_1_and_2()

            # label info1/2/3
            self.save_to_excel()
            self.label_info1.setText("Fall time: {0} [s]\n\n"
                                     "Reflection travel: {1} [m]\n\n"
                                     "Reflection time: {2} [s]\n\n"
                                     "Force from method 1: {3} [N]\n\n"
                                     "Force from method 2: {4} [N]".format(self.fallTime,
                                                                           self.reflectionDistance,
                                                                           self.reflectionTime,
                                                                           self.impactForceMethod1,
                                                                           self.impactForceMethod2))

            # start sending data to creating plots
            # accelerometer 3 axies date
            self.plot_accelerometer(self.fullTimeDate, self.fullAccelerometerXDate, 'b', "Acce. X")
            self.plot_accelerometer(self.fullTimeDate, self.fullAccelerometerYDate, 'g', "Acce. Y")
            self.plot_accelerometer(self.fullTimeDate, self.fullAccelerometerZDate, 'r', "Acce. Z")
            # all sensors combined
            tempAcceX = []
            tempAcceY = []
            tempAcceZ = []
            tempLowerSensor = []
            tempUpperSensor = []
            tempTime1 = []
            for var in range(0, len(self.fullTimeDate)):
                if self.fullUpperSensorDate[var] == 1:
                    tempAcceX.append(self.fullAccelerometerXDate[var])
                    tempAcceY.append(self.fullAccelerometerYDate[var])
                    tempAcceZ.append(self.fullAccelerometerZDate[var])
                    tempLowerSensor.append(self.fullLowerSensorDate[var])
                    tempUpperSensor.append(self.fullUpperSensorDate[var])
                    tempTime1.append(self.fullTimeDate[var])
            # jesli wykresy maja wyswietlic wszystko
            # self.plot_all(self.fullTimeDate, self.fullAccelerometerXDate, 'g', "Acce. X")
            # self.plot_all(self.fullTimeDate, self.fullAccelerometerYDate, 'y', "Acce. Y")
            # self.plot_all(self.fullTimeDate, self.fullAccelerometerZDate, 'c', "Acce. Z")
            # self.plot_all(self.fullTimeDate, self.fullUpperSensorDate, 'b', "S. Upper")
            # self.plot_all(self.fullTimeDate, self.fullLowerSensorDate, 'r', "S. Lower")
            self.plot_all(tempTime1, tempAcceX, 'b', "Acce. X")
            self.plot_all(tempTime1, tempAcceY, 'g', "Acce. Y")
            self.plot_all(tempTime1, tempAcceZ, 'r', "Acce. Z")
            self.plot_all(tempTime1, tempUpperSensor, 'y', "S. Upper")
            self.plot_all(tempTime1, tempLowerSensor, 'c', "S. Lower")
        except Exception as e:
            self.statusBar.showMessage("Plot generating: {}".format(e), 25000)

    def fall_time(self):
        for var in range(0, len(self.fullUpperSensorDate)):
            if self.fullUpperSensorDate[var] == 1:
                self.upperSensor = self.fullTimeDate[var]
                self.upperIndex = var
                break
        for var2 in range(0, len(self.fullLowerSensorDate)):
            if self.fullLowerSensorDate[var2] == 0:
                self.lowerSensor = self.fullTimeDate[var2]
                self.lowerIndex = var2
                break
        self.fallTime = (self.lowerSensor - self.upperSensor)

    def average_acceleration(self):
        temp = 0.0
        count = 0
        for var in range(int(self.upperIndex), int(self.lowerIndex)):
            temp += self.fullAccelerometerXDate[var]
            count += 1
        self.averageAcce = temp / count

    def impact_force_method_1_and_2(self):
        self.average_acceleration()
        self.velocity = math.sqrt(2 * self.averageAcce * self.height)
        self.wnet = 0.5 * self.mass * math.pow(self.velocity, 2)
        self.impactForceMethod1 = self.wnet / self.reflectionDistance
        self.impactForceMethod2 = self.mass * self.reflectionAcceleration

    def reflection_travel(self):
        temp = []
        temp2 = 0.0
        for var in range(0, len(self.fullAccelerometerXDate)):
            if self.fullAccelerometerXDate[var] < 9.0:
                temp.append(self.fullTimeDate[var])
                print(self.fullTimeDate[var])
                temp2 += self.fullAccelerometerXDate[var]

        for var2 in range(1, len(self.fullAccelerometerXDate)):
            if self.fullAccelerometerXDate[var2] < 9.0 and self.fullAccelerometerXDate[var2] < \
                    self.fullAccelerometerXDate[var2 - 1]:
                tempTime = self.fullTimeDate[var2] - self.fullTimeDate[var2 - 1]
                self.reflectionDistance += (abs(tempTime) * abs(self.fullAccelerometerXDate[var2]))

        self.reflectionTime = (temp[len(temp) - 1] - temp[0])
        self.reflectionAcceleration = 2.0 * (
            math.sqrt(2.0 * (temp2/len(temp)) * abs(self.reflectionDistance))) / self.reflectionTime

    def load_from_excel(self):
        wb = load_workbook('wyniki.xlsx')
        print(wb.sheetnames)
        sheet = wb['wyniki2']
        max_rows = sheet.max_row
        for var in range(1, max_rows):
            self.fullTimeDate.append(float(sheet.cell(row=var, column=1).value))
            self.fullUpperSensorDate.append(float(sheet.cell(row=var, column=2).value))
            self.fullLowerSensorDate.append(float(sheet.cell(row=var, column=3).value))
            self.fullAccelerometerXDate.append(float(sheet.cell(row=var, column=4).value))
            self.fullAccelerometerYDate.append(float(sheet.cell(row=var, column=5).value))
            self.fullAccelerometerZDate.append(float(sheet.cell(row=var, column=6).value))

    def save_to_excel(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = datetime.datetime.now()
        worksheet['A2'] = "Time"
        worksheet['B2'] = "Upper Sensor"
        worksheet['C2'] = "Lower Sensor"
        worksheet['D2'] = "Acce X"
        worksheet['E2'] = "Acce Y"
        worksheet['F2'] = "Acce Z"
        worksheet['G2'] = "Reflection Travel: {}".format(self.reflectionDistance)
        worksheet['H2'] = "Reflection Time: {}".format(self.reflectionTime)
        worksheet['I2'] = "Fall time:{}".format(self.fallTime)
        worksheet['J2'] = "Force from distance: {}".format(self.impactForceMethod1)
        worksheet['K2'] = "Force from time: {}".format(self.impactForceMethod2)
        for var in range(0, len(self.fullTimeDate)):
            worksheet.cell(var + 3, 1, self.fullTimeDate[var])
            worksheet.cell(var + 3, 2, self.fullUpperSensorDate[var])
            worksheet.cell(var + 3, 3, self.fullLowerSensorDate[var])
            worksheet.cell(var + 3, 4, self.fullAccelerometerXDate[var])
            worksheet.cell(var + 3, 5, self.fullAccelerometerYDate[var])
            worksheet.cell(var + 3, 6, self.fullAccelerometerZDate[var])
        workbook.save("Wyniki5.xlsx")

    def plot_all(self, time, axis, color, name):
        pen = pg.mkPen(color=color, width=2)
        self.widget_plot_all.setBackground('w')
        self.widget_plot_all.addLegend()
        self.widget_plot_all.setLabel("left", "Acceleration [m/s2]")
        self.widget_plot_all.setLabel("bottom", "Time [s]")
        self.widget_plot_all.showGrid(x=True, y=True)
        self.widget_plot_all.plot(time, axis, pen=pen, name=name)

    def plot_accelerometer(self, time, axis, color, name):
        pen = pg.mkPen(color=color, width=2)
        self.widget_plot_accelerometer.setBackground('w')
        self.widget_plot_accelerometer.addLegend()
        self.widget_plot_accelerometer.setLabel("left", "Acceleration [m/s2]")
        self.widget_plot_accelerometer.setLabel("bottom", "Time [s]")
        self.widget_plot_accelerometer.showGrid(x=True, y=True)
        self.widget_plot_accelerometer.plot(time, axis, pen=pen, name=name)


def main():
    app = QtWidgets.QApplication(sys.argv)
    main = MainWindow()
    main.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
